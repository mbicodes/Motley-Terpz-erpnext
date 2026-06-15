"""
Import CRM Leads from CRM Sales.xlsx (Google Sheet export).

Sheets → pipelines:
  Sales Brand               → Rosin / Solventless
  Sales- Frozen Manufactures → Fresh Frozen
  Sales Retailers            → Retail / Distro

For every imported lead we also attempt to match an existing ERPNext Customer
by name and pull live AR/financial data (same logic as the nightly sync).

Idempotent: rows matched by (lead_name, custom_pipeline) are updated, not duplicated.

Usage:
  bench --site <site> execute cannabis_management.api.import_crm_leads.execute
  bench --site <site> execute cannabis_management.api.import_crm_leads.execute \
      --args '["/path/to/CRM Sales.xlsx"]'
"""

import re
import frappe
from frappe.utils import flt, getdate, nowdate, add_days
from datetime import date, timedelta

DEFAULT_XLSX = "/home/frappeuser/frappe-bench/CRM Sales.xlsx"

# ── Owner map: lowercase first-name → verified ERPNext User ──────────────────

OWNER_MAP = {
    "nikki":  "nikki@motleyterpz.com",
    "matt":   "matt@motleyterpz.com",
    "jamie":  "jamie@motleyterpz.com",
    "imran":  "imran@motleyterpz.com",
    "lizzy":  "lizzy@motleyterpz.com",
    "manny":  "manny@motleyterpz.com",
    "sean":   "sean@motleyterpz.com",
    "tori":   "tori@motleyterpz.com",
}

PIPELINE_COMPANY = {
    "Fresh Frozen":       "TSBC Ranch",
    "Rosin / Solventless":"Motley Terpz",
    "Retail / Distro":    "Both",
    "Tolling":            "Motley Terpz",
}

ACTIVITY_MAP = {
    "consistent":            "Consistent",
    "constistnet":           "Consistent",
    "constistent":           "Consistent",
    "inconsistent":          "Inconsistent",
    "deposit":               "Deposit",
    "have not purchased":    "Never Purchased",
    "no puchase in a year":  "Never Purchased",
    "never purchased":       "Never Purchased",
    "collab":                "Collab",
    "merch":                 "Have not contacted",
    "have not contacted":    "Have not contacted",
    "have not contact":      "Have not contacted",
    # Extra values found in the sheet
    "white label":           "Collab",
    "white labeling":        "Collab",
    "concentrates/ cpg":     "Collab",
    "concentrates/cpg":      "Collab",
    "cpg":                   "Collab",
    "wholesale":             "Consistent",
    "active":                "Consistent",
    "inactive":              "Inconsistent",
}

# Col-A values that indicate a non-data row (only the very specific legend markers)
SKIP_TIER_VALUES = {"relationship status", "click up"}

# col-B values that are sub-headers inside the data area — skip
SKIP_ACCOUNT_NAMES = {
    "account", "account ", "monthly demand", "monthly sales ",
    "monthly sales", "client info", "store", "retailers 2025",
}

BLOCKED_AR    = 50_000.0
BLOCKED_AGING = 90


# ── Entry point ───────────────────────────────────────────────────────────────

def execute(xlsx_path=None):
    try:
        import openpyxl
    except ImportError:
        frappe.throw("openpyxl not found — run: pip install openpyxl")

    path = xlsx_path or DEFAULT_XLSX
    wb   = openpyxl.load_workbook(path, data_only=True)

    sheet_map = {
        "Sales Brand":               "Rosin / Solventless",
        "Sales- Frozen Manufactures":"Fresh Frozen",
        "Sales Retailers":           "Retail / Distro",
    }

    stats = {"created": 0, "updated": 0, "skipped": 0, "errors": 0}

    for sheet_name, pipeline in sheet_map.items():
        if sheet_name not in wb.sheetnames:
            print(f"[import] ⚠  Sheet '{sheet_name}' not found — skipping")
            continue
        ws = wb[sheet_name]
        print(f"\n[import] Processing '{sheet_name}' → {pipeline}")
        if sheet_name == "Sales Retailers":
            _import_retailers(ws, stats)
        else:
            _import_brand_sheet(ws, pipeline, stats)

    frappe.db.commit()
    print(
        f"\n[import] ✓ Done — "
        f"created={stats['created']}  updated={stats['updated']}  "
        f"skipped={stats['skipped']}  errors={stats['errors']}"
    )
    return stats


# ── Sheet parsers ─────────────────────────────────────────────────────────────

def _import_brand_sheet(ws, pipeline, stats):
    """
    Handles 'Sales Brand' and 'Sales- Frozen Manufactures'.
    Header is on row 8; data starts row 9.
    Column positions are read from the header row so the two sheets
    (which differ by one column) are handled identically.
    """
    # Build col index from header row 8
    col = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(8, c).value
        if v:
            col[str(v).strip().upper()] = c

    def g(r, *keys):
        for k in keys:
            idx = col.get(k.upper())
            if idx:
                return ws.cell(r, idx).value
        return None

    company = PIPELINE_COMPANY.get(pipeline, "")

    for r in range(9, ws.max_row + 1):
        tier_raw    = ws.cell(r, 1).value
        account_raw = ws.cell(r, 2).value

        if not account_raw:
            continue
        account_name = str(account_raw).strip()
        if account_name.lower() in SKIP_ACCOUNT_NAMES:
            continue

        parsed = _parse_tier_owner(tier_raw)
        if parsed is None:
            stats["skipped"] += 1
            continue

        # Flags column (col F = "Status" or "Frozen Buyer" depending on sheet)
        flags_raw = g(r, "Status", "Frozen Buyer") or ""
        flags_str = str(flags_raw).strip() if flags_raw else ""

        activity_raw = g(r, "Buyer Activity") or g(r, "BUYER ACTIVITY") or ""
        notes_parts  = []

        n1 = g(r, "Notes")
        n2 = g(r, "What do you know")
        if n1: notes_parts.append(str(n1).strip())
        if n2: notes_parts.append(str(n2).strip())

        # ClickUp
        clickup_raw = g(r, "ClickUp Notes", "CLICKUP NOTES")
        clickup = ""
        if clickup_raw:
            s = str(clickup_raw).strip()
            if s.startswith("http"):
                clickup = s
            elif s and s not in {"????", "-", "n/a"}:
                notes_parts.append(f"ClickUp: {s}")

        notes = "\n".join(filter(None, notes_parts))

        # Demand fields
        demand = {
            "custom_demand_fresh_frozen": _num(g(r, "FROZEN")),
            "custom_demand_rosin":        _num(g(r, "ROSIN")),
            "custom_demand_vrr":          _num(g(r, "VRR")),
            "custom_demand_food_grade":   _num(g(r, "FOOD GRD")),
            "custom_demand_bubble":       _num(g(r, "BUBBLE")),
            "custom_demand_cpg":          _num(g(r, "CPG")),
            "custom_demand_bho":          _num(g(r, "BHO LIVE")) + _num(g(r, "BHO")),
            "custom_demand_flower":       _num(g(r, "FLOWER")) + _num(g(r, "P-ROLLS")),
            "custom_demand_trim":         _num(g(r, "TRIM")),
            "custom_demand_thca":         _num(g(r, "THCA")),
        }

        # Last contact date
        last_contact = _parse_date(g(r, "Last contact", "LAST CONTACT"))

        data = {
            "custom_relationship_tier": parsed["tier"],
            "custom_account_owner":     parsed["owner"] or "",
            "custom_pipeline":          pipeline,
            "custom_company":           company,
            "custom_buyer_activity":    _normalize_activity(activity_raw),
            "custom_last_contact_date": last_contact,
            "custom_notes":             notes,
            "custom_single_source":     1 if "single source" in flags_str.lower() else 0,
            "custom_no_ocal":           1 if "ocal" in flags_str.lower() else 0,
            "custom_clickup_link":      clickup,
            **demand,
        }

        # Account flags
        extra_flags = []
        if "qc" in flags_str.lower() or "custom" in flags_str.lower():
            extra_flags.append("Custom QC Process")
        data["custom_account_flags"] = ", ".join(extra_flags) if extra_flags else ""

        # CRM status from activity + tier
        data["status"] = _infer_status(parsed["tier"], data["custom_buyer_activity"], parsed.get("dead"))

        _upsert_lead(account_name, pipeline, data, stats)


def _import_retailers(ws, stats):
    """
    Handles 'Sales Retailers'.
    Header on row 2 (cols: Store, Owners, Location, Info, Locations, Confirmed, To Big?)
    Data from row 7 (skip rows 3-6 which are category headers / summary counts).
    """
    pipeline = "Retail / Distro"
    company  = "Both"

    for r in range(7, ws.max_row + 1):
        name_raw = ws.cell(r, 1).value
        if not name_raw:
            continue
        account_name = str(name_raw).strip()
        if account_name.lower() in SKIP_ACCOUNT_NAMES | {"to big?", "confirmed ", "confirmed"}:
            continue
        if re.match(r'^[A-Z\s/]+$', account_name) and len(account_name) < 20:
            # likely a section header like "CONFIRMED ", "TO BIG?"
            continue

        owner_raw = ws.cell(r, 2).value
        notes_raw = ws.cell(r, 4).value

        owner = _first_owner(str(owner_raw) if owner_raw else "")
        notes = str(notes_raw).strip() if notes_raw else ""

        data = {
            "custom_relationship_tier": "WIP",
            "custom_account_owner":     owner or "",
            "custom_pipeline":          pipeline,
            "custom_company":           company,
            "custom_buyer_activity":    "Have not contacted",
            "custom_notes":             notes,
            "status":                   "Lead",
        }

        _upsert_lead(account_name, pipeline, data, stats)


# ── Lead upsert ───────────────────────────────────────────────────────────────

BATCH_COMMIT = 50  # commit to DB every N records


def _upsert_lead(account_name, pipeline, data, stats):
    try:
        # Both first_name and organization are capped at 140 chars
        name140 = account_name[:140]

        existing = frappe.db.get_value(
            "CRM Lead",
            {"first_name": name140, "custom_pipeline": pipeline},
            "name",
        )

        if existing:
            doc = frappe.get_doc("CRM Lead", existing)
            for k, v in data.items():
                setattr(doc, k, v)
            doc.save(ignore_permissions=True)
            stats["updated"] += 1
        else:
            doc = frappe.new_doc("CRM Lead")
            doc.first_name   = name140          # mandatory; drives lead_name
            doc.organization = name140
            doc.lead_owner   = data.get("custom_account_owner") or ""
            for k, v in data.items():
                setattr(doc, k, v)
            doc.insert(ignore_permissions=True)
            stats["created"] += 1

        # Batch commit
        total = stats["created"] + stats["updated"]
        if total % BATCH_COMMIT == 0:
            frappe.db.commit()

        # Try to sync live ERP data
        _sync_erp_data(doc.name, account_name)

    except Exception as exc:
        stats["errors"] += 1
        frappe.log_error(frappe.get_traceback(), f"[import_crm_leads] {account_name[:80]}")
        print(f"  [ERROR] {account_name[:80]}: {exc}")


# ── ERP financial sync ────────────────────────────────────────────────────────

def _sync_erp_data(lead_name, account_name):
    """
    Attempt to match the lead to an ERPNext Customer and populate
    all live AR/financial fields exactly as the nightly sync does.
    """
    # Try exact match first, then case-insensitive contains
    customer = frappe.db.get_value("Customer", {"name": account_name}, "name")
    if not customer:
        # Try partial / fuzzy match
        rows = frappe.db.sql(
            "SELECT name FROM `tabCustomer` WHERE name LIKE %s LIMIT 1",
            f"%{account_name[:30]}%",
        )
        if rows:
            customer = rows[0][0]

    if not customer:
        return  # no match — leave ERP fields blank

    data = _pull_erp_data(customer)
    if not data:
        return

    frappe.db.set_value("CRM Lead", lead_name, {
        "custom_erp_customer":        customer,
        "custom_ar_balance":          data["ar_balance"],
        "custom_ar_aging_days":       data["ar_aging_days"],
        "custom_ar_status":           data["ar_status"],
        "custom_cod_flag":            data["cod_flag"],
        "custom_last_invoice_date":   data["last_invoice_date"],
        "custom_last_invoice_amount": data["last_invoice_amount"],
        "custom_last_payment_date":   data["last_payment_date"],
        "custom_mtd_revenue":         data["mtd_revenue"],
        "custom_trailing_8w_revenue": data["trailing_8w_revenue"],
        "custom_payment_terms":       data["payment_terms"],
        "custom_last_sync":           frappe.utils.now(),
    })


def _pull_erp_data(customer):
    today = getdate(nowdate())
    mtd_start  = today.replace(day=1)
    week8_start = today - timedelta(weeks=8)

    ar = frappe.db.sql("""
        SELECT
            COALESCE(SUM(outstanding_amount), 0)            AS balance,
            COALESCE(MAX(DATEDIFF(CURDATE(), due_date)), 0)  AS max_aging
        FROM `tabSales Invoice`
        WHERE customer = %s AND docstatus = 1
          AND outstanding_amount > 0.01
    """, customer, as_dict=True)

    if not ar:
        return None

    balance    = flt(ar[0].balance)
    aging_days = int(ar[0].max_aging or 0)

    if balance > BLOCKED_AR or aging_days > BLOCKED_AGING:
        ar_status = "Blocked"
    elif aging_days > 30:
        ar_status = "Overdue"
    elif balance > 0:
        ar_status = "Watch"
    else:
        ar_status = "Clean"

    inv = frappe.db.sql("""
        SELECT posting_date, grand_total
        FROM `tabSales Invoice`
        WHERE customer = %s AND docstatus = 1
        ORDER BY posting_date DESC LIMIT 1
    """, customer, as_dict=True)

    pay = frappe.db.sql("""
        SELECT MAX(posting_date) AS last_pay
        FROM `tabPayment Entry`
        WHERE party = %s AND party_type = 'Customer'
          AND docstatus = 1 AND payment_type = 'Receive'
    """, customer, as_dict=True)

    mtd = frappe.db.sql("""
        SELECT COALESCE(SUM(grand_total), 0) AS rev
        FROM `tabSales Invoice`
        WHERE customer = %s AND docstatus = 1
          AND posting_date >= %s
    """, (customer, str(mtd_start)), as_dict=True)

    w8 = frappe.db.sql("""
        SELECT COALESCE(SUM(grand_total), 0) AS rev
        FROM `tabSales Invoice`
        WHERE customer = %s AND docstatus = 1
          AND posting_date >= %s
    """, (customer, str(week8_start)), as_dict=True)

    terms = frappe.db.get_value("Customer", customer, "payment_terms") or ""
    cod_flag = 1 if ("cod" in terms.lower() or "cash on delivery" in terms.lower()) else 0

    return {
        "ar_balance":          balance,
        "ar_aging_days":       aging_days,
        "ar_status":           ar_status,
        "cod_flag":            cod_flag,
        "last_invoice_date":   str(inv[0].posting_date)[:10] if inv else None,
        "last_invoice_amount": flt(inv[0].grand_total) if inv else 0,
        "last_payment_date":   str(pay[0].last_pay)[:10] if pay and pay[0].last_pay else None,
        "mtd_revenue":         flt(mtd[0].rev) if mtd else 0,
        "trailing_8w_revenue": flt(w8[0].rev) if w8 else 0,
        "payment_terms":       terms,
    }


# ── Helpers ───────────────────────────────────────────────────────────────────

def _parse_tier_owner(raw):
    """
    Parse col-A value like "1 Nikki AAA", "Matt A", "Jake", "WIP", etc.
    Returns dict, or None only for explicit non-data markers (legend rows, "Click up").
    Empty col A → default tier "Lead".
    """
    if not raw:
        return {"tier": "Lead", "owner": None, "dead": False}
    s = str(raw).strip()
    if not s:
        return {"tier": "Lead", "owner": None, "dead": False}
    if s.lower() in SKIP_TIER_VALUES:
        return None

    sl = s.lower()
    is_dead = "dead" in sl

    # Tier
    if "aaa" in sl:
        tier = "AAA"
    elif "aa" in sl:
        tier = "AA"
    elif re.search(r'\ba\b', sl) or sl.endswith(" a") or (sl.startswith("a ") and "aaa" not in sl):
        tier = "A"
    elif "friends" in sl or "family" in sl:
        tier = "Friends & Family"
    elif "wip" in sl:
        tier = "WIP"
    elif re.search(r'\b1\b', sl):
        tier = "AAA"   # "Matt 1" = Matt's tier 1 accounts
    elif "not in business" in sl:
        tier = "Lead"
    else:
        tier = "Lead"

    # Owner — find first known name
    owner = _first_owner(s)

    return {"tier": tier, "owner": owner, "dead": is_dead}


def _first_owner(raw):
    """Extract the first recognizable owner email from a string."""
    words = re.findall(r"[a-zA-Z']+", raw)
    for w in words:
        email = OWNER_MAP.get(w.lower().rstrip("s"))
        if email:
            return email
        email = OWNER_MAP.get(w.lower())
        if email:
            return email
    return None


VALID_ACTIVITIES = {"Consistent", "Inconsistent", "Deposit", "Never Purchased", "Collab", "Have not contacted"}

def _normalize_activity(raw):
    if not raw:
        return ""
    mapped = ACTIVITY_MAP.get(str(raw).strip().lower())
    if mapped:
        return mapped
    # If the raw value (title-cased) is already a valid option, use it
    title = str(raw).strip().title()
    if title in VALID_ACTIVITIES:
        return title
    return ""  # unknown activity → blank (avoids validation error)


def _infer_status(tier, activity, dead=False):
    if dead:
        return "Inactive"
    if tier in ("Lead",):
        return "Lead"
    if tier == "WIP":
        return "Contacted"
    if activity in ("Consistent", "Inconsistent", "Deposit", "Collab"):
        return "Active"
    if activity == "Never Purchased":
        return "Contacted"
    return "Lead"


def _num(v):
    """Parse a demand value to float lbs. Returns 0.0 for text/None/errors."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        try:
            f = float(v)
            return 0.0 if (f != f) else round(f, 1)  # NaN → 0
        except Exception:
            return 0.0
    s = str(v).strip().lower()
    if not s or s in {"none", "n/a", "-", "#ref!", "#ref", "«"}:
        return 0.0
    # Strip commas
    s = s.replace(",", "")
    m = re.search(r"[\d]+\.?\d*", s)
    if not m:
        return 0.0
    num = float(m.group())
    if "kilo" in s or " kg" in s:
        num = num * 2.20462
    elif s.endswith("k") or " k " in s or "k a month" in s:
        num = num * 1000
    return round(num, 1)


def _parse_date(v):
    if not v:
        return None
    if hasattr(v, "date"):
        return str(v.date())
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y", "%B %d, %Y"):
        try:
            from datetime import datetime as dt
            return dt.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None
