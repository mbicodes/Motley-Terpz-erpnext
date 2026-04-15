# cannabis_management/api/slack_inventory.py
#
# Slack Slash Command handler for /inventory
# Returns an Excel file with inventory data by item group
#
# Endpoint: /api/method/cannabis_management.api.slack_inventory.handle_inventory_command

import frappe
import requests
import json
from io import BytesIO
from datetime import datetime


@frappe.whitelist(allow_guest=True, methods=["POST"])
def handle_inventory_command(**kwargs):
    """
    Handle /inventory slash command from Slack.

    Slack sends form-encoded POST with:
        - token, team_id, team_domain
        - channel_id, channel_name
        - user_id, user_name
        - command, text
        - response_url, trigger_id
    """

    # ── Extract Slack payload ──────────────────────────────────────
    text = frappe.form_dict.get("text", "").strip()
    channel_id = frappe.form_dict.get("channel_id", "")
    user_id = frappe.form_dict.get("user_id", "")
    user_name = frappe.form_dict.get("user_name", "")
    response_url = frappe.form_dict.get("response_url", "")

    # ── Validate this looks like a Slack request ──────────────────
    if not response_url or not channel_id:
        frappe.throw("Invalid request: missing Slack parameters", frappe.AuthenticationError)

    # ── No text = show available item groups ──────────────────────
    if not text:
        frappe.enqueue(
            "cannabis_management.api.slack_inventory.send_item_groups_list",
            response_url=response_url,
            queue="short",
            now=frappe.flags.in_test,
        )
        frappe.response["type"] = "text"
        frappe.response["message"] = ""
        return

    # ── Text provided = generate inventory report ─────────────────
    frappe.enqueue(
        "cannabis_management.api.slack_inventory.generate_and_upload_report",
        item_group_input=text,
        channel_id=channel_id,
        user_id=user_id,
        user_name=user_name,
        response_url=response_url,
        queue="short",
        now=frappe.flags.in_test,
    )

    frappe.clear_messages()
    frappe.response.update({
        "type": "json",
        "http_status_code": 200,
        "response_type": "ephemeral",
        "text": "",
    })
    return


# ─────────────────────────────────────────────────────────────────────
# Background Jobs
# ─────────────────────────────────────────────────────────────────────


def send_item_groups_list(response_url):
    """Send the list of available item groups to Slack."""

    groups = frappe.db.sql(
        """
        SELECT
            ig.name,
            COUNT(i.name) as item_count
        FROM `tabItem Group` ig
        INNER JOIN `tabItem` i ON i.item_group = ig.name 
            AND i.disabled = 0 
            AND i.custom_show_in_dashboard = 1
        WHERE ig.is_group = 0
          AND ig.custom_show_in_dashboard = 1
        GROUP BY ig.name
        ORDER BY ig.name
    """,
        as_dict=True,
    )

    if not groups:
        _slack_respond(response_url, "📭 No item groups are configured for the inventory dashboard.")
        return

    lines = [f"• `{g.name}` ({g.item_count} items)" for g in groups]
    group_list = "\n".join(lines)

    _slack_respond(
        response_url,
        f"📦 *Available Item Groups:*\n\n{group_list}\n\n*Usage:* `/inventory Fresh Frozen`",
    )


def generate_and_upload_report(item_group_input, channel_id, user_id, user_name, response_url):
    """
    Background job:
    1. Validate item group
    2. Query stock data
    3. Generate Excel
    4. Upload to Slack channel using new upload API
    """

    bot_token = frappe.conf.get("slack_bot_token")
    if not bot_token:
        _slack_respond(
            response_url,
            "❌ Slack bot token is not configured. Ask your admin to run:\n"
            "`bench --site <site> set-config slack_bot_token xoxb-YOUR-TOKEN`",
        )
        return

    # ── Case-insensitive item group lookup ────────────────────────
    item_group = _resolve_item_group(item_group_input)

    if not item_group:
        _slack_respond(
            response_url,
            f"❌ Item group *{item_group_input}* not found.\n"
            "Use `/inventory` (with no arguments) to see available groups.",
        )
        return

    # ── Notify user ───────────────────────────────────────────────
    _slack_respond(
        response_url,
        f"⏳ Generating inventory report for *{item_group}*... hang tight.",
    )

    # ── Query stock data ──────────────────────────────────────────
    items = frappe.db.sql(
        """
        SELECT
            b.item_code,
            i.item_name,
            b.warehouse,
            b.actual_qty,
            b.reserved_qty
        FROM `tabBin` b
        INNER JOIN `tabItem` i ON i.name = b.item_code
        WHERE i.item_group = %s
          AND i.disabled = 0
          AND i.custom_show_in_dashboard = 1
          AND b.actual_qty != 0
        ORDER BY i.item_name, b.warehouse
    """,
        (item_group,),
        as_dict=True,
    )

    if not items:
        _slack_respond(response_url, f"📭 No stock found for item group *{item_group}*.")
        return

    # ── Generate Excel ────────────────────────────────────────────
    excel_buffer = _build_excel(items, item_group)
    file_bytes = excel_buffer.getvalue()

    # ── Compute summary for the Slack message ─────────────────────
    total_qty = sum(row.actual_qty or 0 for row in items)
    total_reserved = sum(row.reserved_qty or 0 for row in items)
    total_available = total_qty - total_reserved
    unique_items = len({row.item_code for row in items})
    low_stock = sum(1 for row in items if (row.actual_qty or 0) - (row.reserved_qty or 0) <= 0)

    timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Inventory_{item_group.replace(' ', '_')}_{timestamp_str}.xlsx"

    comment = (
        f"📦 *{item_group} — Inventory Report*\n"
        f"Requested by <@{user_id}>\n"
        f"Generated: {datetime.now().strftime('%b %d, %Y %I:%M %p')}\n\n"
        f"*Unique Items:* {unique_items}  |  "
        f"*Total Qty:* {total_qty:,.2f}  |  "
        f"*Reserved:* {total_reserved:,.2f}  |  "
        f"*Available:* {total_available:,.2f}  |  "
        f"*Low Stock:* {low_stock}"
    )

    # ── Upload to Slack (new V2 API: 3-step process) ──────────────
    try:
        _slack_upload_file_v2(
            bot_token=bot_token,
            channel_id=channel_id,
            file_bytes=file_bytes,
            filename=filename,
            title=f"{item_group} Inventory Report",
            initial_comment=comment,
        )
    except Exception as e:
        error_msg = str(e)
        _slack_respond(
            response_url,
            f"❌ Failed to upload file to Slack: `{error_msg}`\n"
            "Make sure the bot is invited to the channel (`/invite @BotName`).",
        )
        frappe.log_error(
            title="Slack Inventory Upload Failed",
            message=f"Error: {error_msg}",
        )


# ─────────────────────────────────────────────────────────────────────
# Slack Upload V2 (replaces deprecated files.upload)
# ─────────────────────────────────────────────────────────────────────


def _slack_upload_file_v2(bot_token, channel_id, file_bytes, filename, title, initial_comment):
    """
    Upload a file to Slack using the new 3-step process:
    1. files.getUploadURLExternal  → get a presigned upload URL + file_id
    2. POST the file bytes to that URL
    3. files.completeUploadExternal → finalize and share to channel
    """

    headers = {"Authorization": f"Bearer {bot_token}"}

    # ── Step 1: Get upload URL ────────────────────────────────────
    step1_resp = requests.get(
        "https://slack.com/api/files.getUploadURLExternal",
        headers=headers,
        params={
            "filename": filename,
            "length": len(file_bytes),
        },
        timeout=15,
    )
    step1_data = step1_resp.json()

    if not step1_data.get("ok"):
        raise Exception(f"getUploadURLExternal failed: {step1_data.get('error', 'unknown')}")

    upload_url = step1_data["upload_url"]
    file_id = step1_data["file_id"]

    # ── Step 2: Upload file bytes to the presigned URL ────────────
    step2_resp = requests.post(
        upload_url,
        data=file_bytes,
        headers={"Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        timeout=30,
    )

    if step2_resp.status_code != 200:
        raise Exception(f"File upload to presigned URL failed: HTTP {step2_resp.status_code}")

    # ── Step 3: Complete the upload and share to channel ──────────
    step3_resp = requests.post(
        "https://slack.com/api/files.completeUploadExternal",
        headers={**headers, "Content-Type": "application/json"},
        json={
            "files": [{"id": file_id, "title": title}],
            "channel_id": channel_id,
            "initial_comment": initial_comment,
        },
        timeout=15,
    )
    step3_data = step3_resp.json()

    if not step3_data.get("ok"):
        raise Exception(f"completeUploadExternal failed: {step3_data.get('error', 'unknown')}")

    return step3_data


# ─────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────


def _resolve_item_group(user_input):
    """
    Case-insensitive lookup for item group.
    First tries exact match, then LIKE match.
    Returns the canonical Item Group name or None.
    """
    # Exact match first
    if frappe.db.exists("Item Group", user_input):
        return user_input

    # Case-insensitive match
    result = frappe.db.sql(
        "SELECT name FROM `tabItem Group` WHERE LOWER(name) = LOWER(%s) LIMIT 1",
        (user_input,),
    )
    if result:
        return result[0][0]

    # Partial / fuzzy match (contains)
    result = frappe.db.sql(
        "SELECT name FROM `tabItem Group` WHERE LOWER(name) LIKE %s LIMIT 1",
        (f"%{user_input.lower()}%",),
    )
    if result:
        return result[0][0]

    return None


def _slack_respond(response_url, text, response_type="ephemeral"):
    """Send a message back to Slack via the response_url."""
    try:
        requests.post(
            response_url,
            json={"response_type": response_type, "text": text},
            timeout=10,
        )
    except Exception as e:
        frappe.log_error(title="Slack Response Failed", message=str(e))


def _build_excel(items, item_group):
    """Generate a professionally formatted Excel workbook and return a BytesIO buffer."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = item_group[:31]  # Sheet name max 31 chars

    # ── Styles ────────────────────────────────────────────────────
    title_font = Font(name="Calibri", bold=True, size=16, color="1B5E20")
    subtitle_font = Font(name="Calibri", italic=True, size=10, color="757575")
    summary_font = Font(name="Calibri", bold=True, size=11, color="1565C0")

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")

    data_font = Font(name="Calibri", size=10)
    number_font = Font(name="Calibri", size=10)

    warning_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="BDBDBD"),
        right=Side(style="thin", color="BDBDBD"),
        top=Side(style="thin", color="BDBDBD"),
        bottom=Side(style="thin", color="BDBDBD"),
    )

    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    # ── Title Section ─────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = f"{item_group} — Inventory Report"
    cell.font = title_font
    cell.alignment = center
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:F2")
    cell = ws["A2"]
    cell.value = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    cell.font = subtitle_font
    cell.alignment = center
    ws.row_dimensions[2].height = 20

    # ── Summary Row ───────────────────────────────────────────────
    total_qty = sum(row.actual_qty or 0 for row in items)
    total_reserved = sum(row.reserved_qty or 0 for row in items)
    total_available = total_qty - total_reserved
    unique_items = len({row.item_code for row in items})
    low_stock = sum(1 for row in items if (row.actual_qty or 0) - (row.reserved_qty or 0) <= 0)

    ws.merge_cells("A4:F4")
    cell = ws["A4"]
    cell.value = (
        f"Items: {unique_items}   |   "
        f"Total Qty: {total_qty:,.2f}   |   "
        f"Reserved: {total_reserved:,.2f}   |   "
        f"Available: {total_available:,.2f}   |   "
        f"Low Stock: {low_stock}"
    )
    cell.font = summary_font
    cell.alignment = center
    ws.row_dimensions[4].height = 25

    # ── Table Headers (Row 6) ─────────────────────────────────────
    headers = [
        ("Item Code", 25),
        ("Item Name", 40),
        ("Warehouse", 30),
        ("Balance Qty", 15),
        ("Reserved Qty", 15),
        ("Available Qty", 15),
    ]

    header_row = 6
    for col_idx, (header_text, col_width) in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col_idx)].width = col_width

    ws.row_dimensions[header_row].height = 28

    # ── Data Rows ─────────────────────────────────────────────────
    for row_idx, item in enumerate(items, header_row + 1):
        actual = round(item.actual_qty or 0, 2)
        reserved = round(item.reserved_qty or 0, 2)
        available = round(actual - reserved, 2)

        row_data = [
            item.item_code,
            item.item_name or "-",
            item.warehouse or "-",
            actual,
            reserved,
            available,
        ]

        is_low_stock = available <= 0
        is_alt_row = (row_idx - header_row) % 2 == 0

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            if col_idx <= 3:
                cell.font = data_font
                cell.alignment = left_align
            else:
                cell.font = number_font
                cell.alignment = right_align
                cell.number_format = "#,##0.00"

            if is_low_stock:
                cell.fill = warning_fill
            elif is_alt_row:
                cell.fill = alt_row_fill

    # ── Freeze panes ──────────────────────────────────────────────
    ws.freeze_panes = f"A{header_row + 1}"

    # ── Auto-filter ───────────────────────────────────────────────
    last_data_row = header_row + len(items)
    ws.auto_filter.ref = f"A{header_row}:F{last_data_row}"

    # ── Save to buffer ────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf