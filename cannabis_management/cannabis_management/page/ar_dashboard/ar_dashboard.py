import frappe
from contextlib import contextmanager
from frappe.utils import nowdate


RECON_EDIT_ROLES    = ("Account Manager", "System Manager", "Administrator")

# Roles whose holders see org-wide AR numbers (all companies), identical to
# what Administrator sees. Without this, Company User Permissions silently
# zero out non-permitted companies (e.g. matt@motleyterpz.com has Company
# permissions for only TSBC Ranch / Motley Terpz / TMM Group, so Master Touch
# Manufacturing, MTPZ and LA Canna invoices vanish from his totals)edw.
ORG_WIDE_VIEW_ROLES = ("Account Receivable", "System Manager", "CEO")


@contextmanager
def _org_wide_view():
    """Temporarily elevate to Administrator so the AR report ignores
    Company User Permissions for users holding an ORG_WIDE_VIEW_ROLES role.

    DO NOT REMOVE: get_ar_data must stay wrapped in this context manager.
    Anything permission-sensitive for the real user (e.g. _can_edit_recon)
    must be computed BEFORE entering it.
    """
    user = frappe.session.user
    if user == "Administrator" or not set(frappe.get_roles(user)).intersection(ORG_WIDE_VIEW_ROLES):
        yield
        return
    try:
        frappe.set_user("Administrator")
        yield
    finally:
        frappe.set_user(user)
TMM_GROUP_COMPANIES = ["Motley Terpz", "TSBC Ranch"]
LEGACY_CUTOFF       = "2026-05-31"
NEW_AR_START        = "2026-06-01"


ALLOWED_RECON_STATUSES = (
    "",
    "Reconciled collecting money",
    "Reconciled trouble collecting money",
    "Unreconciled",
    "Dispute",
    "Adjustment",
)

# Yes/No columns edited inline on the dashboard. Key = the name the row payload
# and the front-end use; value = the Customer custom field it is stored in.
# Both are plain Selects rather than Checks so "not answered yet" stays
# distinguishable from "No" — a blank cell means nobody has said either way.
AR_FLAG_FIELDS = {
    "onboarding": "custom_onboarding",
    "company_made_contact": "custom_company_made_contact",
}

ALLOWED_FLAG_VALUES = ("", "Yes", "No")


# Legacy fallback — customer records that predate is_internal_customer/represents_company
# being set consistently. Keep in sync with cannabis_management.api.jamie.INTERNAL_CUSTOMERS.
_INTERNAL_CUSTOMER_FALLBACK = ("Motley Terpz", "MT", "MTPZ")


def _internal_customer_names():
    """Return every customer name that represents an in-house entity rather than a
    real external buyer, so intercompany transfers never show up as AR:
      1. Every Company name (exact match)
      2. Every Customer flagged is_internal_customer = 1
      3. Every Customer linked to a company via represents_company
      4. The legacy fallback list (e.g. "MTPZ" — a real Customer row with
         is_internal_customer left at 0, so bucket 2 alone misses it)
    """
    company_names = frappe.db.sql_list("SELECT name FROM `tabCompany`")
    internal = frappe.db.sql_list("""
        SELECT name FROM `tabCustomer`
        WHERE is_internal_customer = 1
           OR (represents_company IS NOT NULL AND represents_company != '')
    """)
    return set(company_names) | set(internal) | set(_INTERNAL_CUSTOMER_FALLBACK)


def _can_edit_recon():
    user_roles = set(frappe.get_roles(frappe.session.user))
    return any(role in user_roles for role in RECON_EDIT_ROLES)


def _build_ranges(range_str):
    range_numbers = [int(r.strip()) for r in range_str.split(",") if r.strip().isdigit()]
    ranges = []
    prev = 0
    for num in range_numbers:
        ranges.append({"key": f"range{len(ranges) + 1}", "label": f"{prev}-{num}"})
        prev = num
    ranges.append({"key": f"range{len(ranges) + 1}", "label": f"{prev}+"})
    return ranges


def _compute_totals(rows, ranges):
    totals = {"invoiced": 0.0, "paid": 0.0, "outstanding": 0.0}
    for r in ranges:
        totals[r["key"]] = 0.0
    for row in rows:
        totals["invoiced"]    += row["invoiced"]
        totals["paid"]        += row["paid"]
        totals["outstanding"] += row["outstanding"]
        for r in ranges:
            totals[r["key"]] += row[r["key"]]
    return totals


def _strip_cross_company_rows(rows, company):
    """Remove Sales Invoice rows whose company field doesn't match the selected company."""
    si_nos = [r["voucher_no"] for r in rows if r.get("voucher_type") == "Sales Invoice"]
    if not si_nos:
        return rows

    valid = set(frappe.db.sql_list(
        "SELECT name FROM `tabSales Invoice` WHERE name IN %(names)s AND company = %(company)s",
        {"names": tuple(si_nos), "company": company},
    ))

    return [
        r for r in rows
        if r.get("voucher_type") != "Sales Invoice" or r["voucher_no"] in valid
    ]


def _apply_si_outstanding(rows, ranges, report_date):
    """
    Replace the GL-derived outstanding AND aging bucket values with figures
    derived from si.outstanding_amount / si.due_date for every Sales Invoice row,
    then drop rows that are fully paid.

    The ERPNext AR report reads from GL Entry, which diverges from
    si.outstanding_amount when payments are applied without full GL clearance.
    si.outstanding_amount is the authoritative paid/unpaid signal.
    """
    si_nos = [r["voucher_no"] for r in rows if r.get("voucher_type") == "Sales Invoice"]
    if not si_nos:
        return rows

    si_data = frappe.db.sql(
        "SELECT name, outstanding_amount, due_date FROM `tabSales Invoice` WHERE name IN %(names)s",
        {"names": tuple(si_nos)},
        as_dict=True,
    )
    si_map = {r.name: r for r in si_data}

    from frappe.utils import date_diff, getdate
    as_of = getdate(report_date or nowdate())

    # Build aging thresholds from ranges: range1 covers 0..limits[0], range2 limits[0]..limits[1], …
    # _build_ranges("30,60,90,120") → [0-30, 30-60, 60-90, 90-120, 120+]
    # We derive limits by parsing each label.
    range_keys = [r["key"] for r in ranges]

    def _bucket_key(days_overdue):
        """Return the range key for a given days-overdue value (0 = current/not yet due)."""
        prev = 0
        for i, rng in enumerate(ranges):
            label = rng["label"]
            if "+" in label:
                return rng["key"]   # last bucket catches everything remaining
            try:
                _, upper = label.split("-")
                upper = int(upper)
            except Exception:
                return rng["key"]
            if days_overdue <= upper:
                return rng["key"]
            prev = upper
        return range_keys[-1]

    result = []
    for r in rows:
        if r.get("voucher_type") == "Sales Invoice":
            si = si_map.get(r["voucher_no"])
            if not si:
                result.append(r)
                continue
            actual = float(si.outstanding_amount or 0)
            if actual <= 0.01:
                continue    # fully paid — exclude

            r = dict(r)
            r["outstanding"] = actual

            # Recalculate aging bucket: zero all ranges, then put actual in correct bucket
            due = getdate(str(si.due_date)) if si.due_date else as_of
            days_overdue = max(date_diff(as_of, due), 0)
            for key in range_keys:
                r[key] = 0.0
            r[_bucket_key(days_overdue)] = actual

        result.append(r)
    return result


def _fetch_rows_for_company(company, report_date, customer, ageing_based_on, range_str, ranges):
    from erpnext.accounts.report.accounts_receivable.accounts_receivable import execute as ar_execute

    filters = frappe._dict({
        "company": company,
        "report_date": report_date or nowdate(),
        "party_type": "Customer",
        "ageing_based_on": ageing_based_on,
        "range": range_str,
    })
    if customer:
        filters["party"] = [customer]

    columns, data, _, _, _, _ = ar_execute(filters)

    rows = []
    for row in (data or []):
        if not row or not row.get("voucher_no"):
            continue
        if row.get("voucher_type") != "Sales Invoice":
            continue

        processed = {
            "party": row.get("party") or "",
            "customer_name": row.get("customer_name") or row.get("party") or "",
            "voucher_type": row.get("voucher_type") or "",
            "voucher_no": row.get("voucher_no") or "",
            "posting_date": str(row.get("posting_date") or ""),
            "due_date": str(row.get("due_date") or ""),
            "invoiced": float(row.get("invoiced") or 0),
            "paid": float(row.get("paid") or 0),
            "outstanding": float(row.get("outstanding") or 0),
            "currency": row.get("currency") or "",
        }
        for r in ranges:
            processed[r["key"]] = float(row.get(r["key"]) or 0)

        rows.append(processed)

    return rows


@frappe.whitelist()
def init_page():
    companies = frappe.get_all("Company", pluck="name", order_by="name")
    if "TMM Group" not in companies:
        companies.append("TMM Group")
    companies.sort()
    return {
        "companies": companies,
        "can_edit_recon": _can_edit_recon(),
    }


@frappe.whitelist()
def get_ar_data(company, report_date=None, customer=None, ageing_based_on="Due Date",
                range_str="30, 60, 90, 120", ar_mode="legacy"):
    """Whitelisted entry point. Computes user-specific flags first, then runs
    the whole report fetch inside _org_wide_view() so dashboard-role users
    (Matt/CEO etc.) always see the same org-wide numbers as Administrator.

    DO NOT edit report logic here — edit _get_ar_data below. This wrapper must
    stay intact so future changes cannot accidentally drop the elevation.
    """
    can_edit_recon = _can_edit_recon()
    with _org_wide_view():
        result = _get_ar_data(company, report_date, customer, ageing_based_on,
                              range_str, ar_mode)
    result["can_edit_recon"] = can_edit_recon
    return result


def _get_ar_data(company, report_date=None, customer=None, ageing_based_on="Due Date",
                 range_str="30, 60, 90, 120", ar_mode="legacy"):
    ranges = _build_ranges(range_str)

    # Aging "as of" date. Legacy aging is computed relative to TODAY (not clamped
    # to the May-31 cutover) — LEGACY_CUTOFF is only used below to classify which
    # invoices are legacy (posting_date <= cutoff), never to freeze the aging date.
    if ar_mode == "legacy":
        if not report_date:
            report_date = nowdate()
    elif ar_mode == "all":
        if not report_date:
            report_date = nowdate()
    else:
        if not report_date or str(report_date) < NEW_AR_START:
            report_date = nowdate()

    if company == "TMM Group":
        all_rows = []
        for c in TMM_GROUP_COMPANIES:
            try:
                c_rows = _fetch_rows_for_company(
                    c, report_date, customer, ageing_based_on, range_str, ranges
                )
                c_rows = _strip_cross_company_rows(c_rows, c)
                c_rows = _apply_si_outstanding(c_rows, ranges, report_date)
                all_rows.extend(c_rows)
            except Exception:
                frappe.log_error(f"AR data fetch failed for {c}", "TMM Group AR Dashboard")
        rows = all_rows
    else:
        rows = _fetch_rows_for_company(
            company, report_date, customer, ageing_based_on, range_str, ranges
        )
        rows = _strip_cross_company_rows(rows, company)
        rows = _apply_si_outstanding(rows, ranges, report_date)

    # Apply mode date filter and strip intercompany rows in one pass
    internal = _internal_customer_names()
    if ar_mode == "legacy":
        rows = [r for r in rows
                if str(r.get("posting_date") or "") <= LEGACY_CUTOFF
                and r.get("party") not in internal]
    elif ar_mode == "all":
        rows = [r for r in rows
                if r.get("party") not in internal]
    else:
        # New AR: keep new invoices as-is, AND keep legacy invoices so legacy
        # customers appear in the New AR view too — but flag them and zero their
        # aging buckets so the legacy amount never enters the aging/term breakdown.
        kept = []
        for r in rows:
            if r.get("party") in internal:
                continue
            posting = str(r.get("posting_date") or "")
            if posting >= NEW_AR_START:
                kept.append(r)
            elif posting <= LEGACY_CUTOFF:
                lr = dict(r)
                lr["is_legacy"] = 1
                for rng in ranges:
                    lr[rng["key"]] = 0.0
                kept.append(lr)
        rows = kept

    # In modes that show the on-terms (0-10 / 10-20 / 20-30) breakdown, an invoice
    # that is NOT yet past its due date belongs only in those term columns — clear
    # its overdue aging buckets so the same amount never also appears in the red
    # 0-30/30-60/... columns. Once the due date passes, the amount drops out of the
    # term columns and shows in the overdue buckets. (Legacy mode keeps standard
    # ERPNext aging, where the first bucket includes current + 0-30 overdue.)
    if ar_mode in ("new", "all"):
        from frappe.utils import getdate
        as_of = getdate(report_date)
        for r in rows:
            due = r.get("due_date")
            not_overdue = (not due) or getdate(str(due)) >= as_of
            if not_overdue:
                for rng in ranges:
                    r[rng["key"]] = 0.0

    totals = _compute_totals(rows, ranges)

    # Attach reconciliation status, the AR flags and new_ar_available from Customer master
    unique_parties = list({r["party"] for r in rows if r.get("party")})
    blank_info = {"recon": "", "new_ar_available": False, "notebox": ""}
    blank_info.update({key: "" for key in AR_FLAG_FIELDS})
    cust_info = {}
    if unique_parties:
        fetch_fields = ["name", "custom_reconciliation_status", "custom_notebox"]
        # The AR flags and custom_new_ar_available are added via
        # setup_ar_custom_fields; fetch them silently if they exist.
        optional = list(AR_FLAG_FIELDS.values()) + ["custom_new_ar_available"]
        try:
            cust_rows = frappe.get_all(
                "Customer",
                filters={"name": ["in", unique_parties]},
                fields=fetch_fields + optional,
            )
        except Exception:
            cust_rows = frappe.get_all(
                "Customer",
                filters={"name": ["in", unique_parties]},
                fields=fetch_fields,
            )
        for c in cust_rows:
            info = {
                "recon": c.get("custom_reconciliation_status") or "",
                "new_ar_available": bool(c.get("custom_new_ar_available")),
                "notebox": c.get("custom_notebox") or "",
            }
            for key, fieldname in AR_FLAG_FIELDS.items():
                info[key] = c.get(fieldname) or ""
            cust_info[c["name"]] = info

    for r in rows:
        info = cust_info.get(r["party"], blank_info)
        r["reconciliation_status"] = info["recon"]
        r["new_ar_available"]      = info["new_ar_available"]
        r["notebox"]               = info.get("notebox", "")
        for key in AR_FLAG_FIELDS:
            r[key] = info.get(key, "")

    return {
        "rows": rows,
        "ranges": ranges,
        "totals": totals,
        "company": company,
        "report_date": str(report_date or nowdate()),
        "can_edit_recon": _can_edit_recon(),
        "ar_mode": ar_mode,
    }


@frappe.whitelist()
def update_recon_status(party, status):
    """Update custom_reconciliation_status on the Customer master.
    Restricted to users with the Account Manager role (plus System Manager / Administrator).
    """
    if not _can_edit_recon():
        frappe.throw(
            "You do not have permission to change the reconciliation status. "
            "This action is restricted to Account Managers.",
            frappe.PermissionError,
        )

    if status not in ALLOWED_RECON_STATUSES:
        frappe.throw("Invalid reconciliation status")

    if not frappe.db.exists("Customer", party):
        frappe.throw(f"Customer {party} not found")

    frappe.db.set_value("Customer", party, "custom_reconciliation_status", status)

    return {"party": party, "status": status}


@frappe.whitelist()
def update_ar_flag(party, field, value):
    """Set one of the inline Yes/No columns (Onboarding, Company Made Contact).

    `field` is the payload key, not a fieldname — it is looked up in
    AR_FLAG_FIELDS so a caller can never steer this at an arbitrary column.
    """
    if not _can_edit_recon():
        frappe.throw(
            "You do not have permission to change this field.",
            frappe.PermissionError,
        )
    fieldname = AR_FLAG_FIELDS.get(field)
    if not fieldname:
        frappe.throw("Unknown field")
    if value not in ALLOWED_FLAG_VALUES:
        frappe.throw("Invalid value — expected Yes or No")
    if not frappe.db.exists("Customer", party):
        frappe.throw(f"Customer {party} not found")
    frappe.db.set_value("Customer", party, fieldname, value)
    return {"party": party, "field": field, "value": value}


@frappe.whitelist()
def update_notebox(party, value):
    """Save free-text notes to the custom_notebox field on the Customer master.
    The note persists until changed, and rides along in every export."""
    if not _can_edit_recon():
        frappe.throw(
            "You do not have permission to edit notes.",
            frappe.PermissionError,
        )
    if not frappe.db.exists("Customer", party):
        frappe.throw(f"Customer {party} not found")

    frappe.db.set_value("Customer", party, "custom_notebox", value or "")
    return {"party": party, "value": value or ""}


def _company_condition(alias, company):
    """Build a (sql_fragment, extra_params) pair scoping `alias`.company to the
    selected filter, matching the page's existing All Entities / TMM Group /
    single-company semantics. Returns ("", {}) for All Entities."""
    if not company or company == "__ALL__":
        return "", {}
    if company == "TMM Group":
        return f"AND {alias}.company IN %(tmm_companies)s", {"tmm_companies": tuple(TMM_GROUP_COMPANIES)}
    return f"AND {alias}.company = %(company)s", {"company": company}


def _month_labels(start_month, end_month):
    """['2026-01', '2026-02', ...] inclusive, from start_month through end_month
    (both 'YYYY-MM' strings)."""
    y, m = (int(x) for x in start_month.split("-"))
    ey, em = (int(x) for x in end_month.split("-"))
    months = []
    while (y, m) <= (ey, em):
        months.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            m = 1
            y += 1
    return months


def _trailing_weekly_blocks_since(start_date, as_of, weeks=4):
    """Fixed 7-day blocks anchored at start_date (e.g. June 1), returning only
    the most recent `weeks` blocks up to as_of. The final block is clipped to
    as_of so a partial current week still shows real numbers, not padding."""
    from frappe.utils import add_days, getdate

    start_date = getdate(start_date)
    as_of = getdate(as_of)

    blocks = []
    cursor = start_date
    while cursor <= as_of:
        block_end = min(add_days(cursor, 6), as_of)
        blocks.append({
            "label": f"{cursor.strftime('%b %d')}–{block_end.strftime('%b %d')}",
            "from_date": str(cursor),
            "to_date": str(block_end),
        })
        cursor = add_days(cursor, 7)

    return blocks[-weeks:]


@frappe.whitelist()
def get_monthly_ar_collection(company=None):
    """Whitelisted entry point — see _get_ar_data's docstring for why this
    must stay wrapped in _org_wide_view(). Do not edit report logic here."""
    with _org_wide_view():
        return _get_monthly_ar_collection(company)


def _get_monthly_ar_collection(company=None):
    """Month-wise cash collected against Legacy AR (invoices posted on/before
    2026-05-31) vs New AR (invoices posted on/after 2026-06-01), split by
    Cash vs Bank (by the receiving account's account_type — same convention
    used by the Sales Overview / Weekly Cash Ledger pages), plus the Legacy
    AR balance at each month end (which can only shrink — no new legacy
    invoices are ever created after the cutoff).

    Only real Payment Entries (money actually received) count here — no
    credit notes / write-offs / journal adjustments.

    Intercompany customers are excluded from every query below.
    """
    internal = _internal_customer_names()
    internal_tuple = tuple(internal) if internal else ("__none__",)

    company_cond, company_params = _company_condition("pe", company)
    params = {
        "internal": internal_tuple,
        "legacy_cutoff": LEGACY_CUTOFF,
        **company_params,
    }

    # ── Cash received: Payment Entry allocations against Sales Invoices ───────
    # receipt_type is Cash only when the receiving account is flagged
    # account_type='Cash' in the Chart of Accounts; everything else is Bank.
    collection_rows = frappe.db.sql(f"""
        SELECT
            DATE_FORMAT(pe.posting_date, '%%Y-%%m') AS month,
            CASE WHEN si.posting_date <= %(legacy_cutoff)s THEN 'legacy' ELSE 'new' END AS ar_type,
            CASE WHEN acc.account_type = 'Cash' THEN 'cash' ELSE 'bank' END AS receipt_type,
            SUM(per.allocated_amount) AS amount
        FROM `tabPayment Entry Reference` per
        JOIN `tabPayment Entry` pe ON pe.name = per.parent
        JOIN `tabSales Invoice` si ON si.name = per.reference_name
        LEFT JOIN `tabAccount` acc ON acc.name = pe.paid_to
        WHERE pe.docstatus = 1
          AND pe.payment_type = 'Receive'
          AND per.reference_doctype = 'Sales Invoice'
          AND si.customer NOT IN %(internal)s
          {company_cond}
        GROUP BY month, ar_type, receipt_type
    """, params, as_dict=True)

    # ── Legacy total invoiced (fixed) — anchor for the shrinking balance ──────
    si_company_cond, si_company_params = _company_condition("si", company)
    si_params = {"internal": internal_tuple, "legacy_cutoff": LEGACY_CUTOFF, **si_company_params}
    legacy_total_invoiced = frappe.db.sql(f"""
        SELECT COALESCE(SUM(si.grand_total), 0)
        FROM `tabSales Invoice` si
        WHERE si.docstatus = 1
          AND si.posting_date <= %(legacy_cutoff)s
          AND si.customer NOT IN %(internal)s
          {si_company_cond}
    """, si_params)[0][0]

    # ── Assemble month grid ────────────────────────────────────────────────────
    today = nowdate()
    all_months = sorted({r.month for r in collection_rows})
    start_month = min(all_months) if all_months else today[:7]
    end_month = today[:7]
    months = _month_labels(start_month, end_month)

    by_month = {
        m: {"legacy_cash": 0.0, "legacy_bank": 0.0, "new_cash": 0.0, "new_bank": 0.0}
        for m in months
    }
    for r in collection_rows:
        if r.month in by_month:
            by_month[r.month][f"{r.ar_type}_{r.receipt_type}"] += float(r.amount or 0)

    running_collected = 0.0
    monthly_rows = []
    for m in months:
        row = by_month[m]
        legacy_total = row["legacy_cash"] + row["legacy_bank"]
        running_collected += legacy_total
        monthly_rows.append({
            "month": m,
            "label": frappe.utils.formatdate(f"{m}-01", "MMM yyyy"),
            "legacy_cash": round(row["legacy_cash"], 2),
            "legacy_bank": round(row["legacy_bank"], 2),
            "legacy_balance": round(legacy_total_invoiced - running_collected, 2),
            "new_cash": round(row["new_cash"], 2),
            "new_bank": round(row["new_bank"], 2),
        })

    # ── Trailing 4 weekly blocks since June 1 (weekly granularity) ────────────
    weekly_blocks = _trailing_weekly_blocks_since(NEW_AR_START, today, weeks=4)
    weekly_rows = []
    if weekly_blocks:
        min_wk, max_wk = weekly_blocks[0]["from_date"], weekly_blocks[-1]["to_date"]
        wk_params = {**params, "wk_start": min_wk, "wk_end": max_wk}
        wk_rows = frappe.db.sql(f"""
            SELECT
                pe.posting_date AS posting_date,
                CASE WHEN si.posting_date <= %(legacy_cutoff)s THEN 'legacy' ELSE 'new' END AS ar_type,
                CASE WHEN acc.account_type = 'Cash' THEN 'cash' ELSE 'bank' END AS receipt_type,
                per.allocated_amount AS amount
            FROM `tabPayment Entry Reference` per
            JOIN `tabPayment Entry` pe ON pe.name = per.parent
            JOIN `tabSales Invoice` si ON si.name = per.reference_name
            LEFT JOIN `tabAccount` acc ON acc.name = pe.paid_to
            WHERE pe.docstatus = 1
              AND pe.payment_type = 'Receive'
              AND per.reference_doctype = 'Sales Invoice'
              AND si.customer NOT IN %(internal)s
              AND pe.posting_date BETWEEN %(wk_start)s AND %(wk_end)s
              {company_cond}
        """, wk_params, as_dict=True)

        from frappe.utils import getdate

        for blk in weekly_blocks:
            fd, td = getdate(blk["from_date"]), getdate(blk["to_date"])
            legacy_cash = sum(float(r.amount or 0) for r in wk_rows
                               if r.ar_type == "legacy" and r.receipt_type == "cash"
                               and fd <= getdate(str(r.posting_date)) <= td)
            legacy_bank = sum(float(r.amount or 0) for r in wk_rows
                               if r.ar_type == "legacy" and r.receipt_type == "bank"
                               and fd <= getdate(str(r.posting_date)) <= td)
            new_cash = sum(float(r.amount or 0) for r in wk_rows
                           if r.ar_type == "new" and r.receipt_type == "cash"
                           and fd <= getdate(str(r.posting_date)) <= td)
            new_bank = sum(float(r.amount or 0) for r in wk_rows
                           if r.ar_type == "new" and r.receipt_type == "bank"
                           and fd <= getdate(str(r.posting_date)) <= td)
            weekly_rows.append({
                "label": blk["label"],
                "from_date": blk["from_date"],
                "to_date": blk["to_date"],
                "legacy_cash": round(legacy_cash, 2),
                "legacy_bank": round(legacy_bank, 2),
                "new_cash": round(new_cash, 2),
                "new_bank": round(new_bank, 2),
            })

    return {
        "company": company,
        "legacy_total_invoiced": round(float(legacy_total_invoiced or 0), 2),
        "monthly": monthly_rows,
        "weekly": weekly_rows,
        "totals": {
            "legacy_cash": round(sum(r["legacy_cash"] for r in monthly_rows), 2),
            "legacy_bank": round(sum(r["legacy_bank"] for r in monthly_rows), 2),
            "new_cash": round(sum(r["new_cash"] for r in monthly_rows), 2),
            "new_bank": round(sum(r["new_bank"] for r in monthly_rows), 2),
            "legacy_balance_now": monthly_rows[-1]["legacy_balance"] if monthly_rows else round(float(legacy_total_invoiced or 0), 2),
        },
    }


@frappe.whitelist()
def update_new_ar_available(party, value):
    """Toggle the custom_new_ar_available flag on the Customer master."""
    if not _can_edit_recon():
        frappe.throw(
            "You do not have permission to change New AR status.",
            frappe.PermissionError,
        )
    if not frappe.db.exists("Customer", party):
        frappe.throw(f"Customer {party} not found")

    frappe.db.set_value("Customer", party, "custom_new_ar_available", int(value))
    return {"party": party, "value": int(value)}


@frappe.whitelist()
def export_ar_excel_data(rows_json, filename):
    """Generate a multi-sheet .xlsx: one 'All' sheet + one sheet per recon status that has data."""
    import json
    import base64
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    all_rows = json.loads(rows_json)
    if not all_rows:
        frappe.throw("No data to export")

    # rows_json = [header, ...data..., TOTAL row]
    header    = all_rows[0]
    data_rows = all_rows[1:]   # keep the TOTAL row — JS already built it correctly

    # Sheet definitions: (recon_status_value, sheet_name, tab_color_hex)
    SHEETS = [
        ("",                                    "No Status",               "9CA3AF"),  # grey
        ("Reconciled collecting money",         "Reconciled (Collecting)", "16A34A"),  # green
        ("Reconciled trouble collecting money", "Reconciled (Trouble)",    "D97706"),  # amber
        ("Unreconciled",                        "Unreconciled",            "DC2626"),  # red
        ("Dispute",                             "Dispute",                 "7C3AED"),  # purple
        ("Adjustment",                          "Adjustment",              "3B82F6"),  # blue
    ]

    # Shared styles
    thin         = Side(style="thin", color="D1D5DB")
    border       = Border(top=thin, bottom=thin, left=thin, right=thin)
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1D4ED8")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    total_font   = Font(bold=True, size=10)
    total_fill   = PatternFill("solid", fgColor="DBEAFE")

    def write_sheet(ws, sheet_data, tab_color=None):
        if tab_color:
            ws.sheet_properties.tabColor = tab_color

        # Header row
        ws.append(header)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        # Data rows — last row is the TOTAL row already built by JS
        for r_idx, row in enumerate(sheet_data, 2):
            is_total = str(row[0] or "").upper() == "TOTAL"
            ws.append(row)
            for c_idx, cell in enumerate(ws[r_idx], 1):
                cell.border = border
                is_num = isinstance(cell.value, (int, float))
                if is_total:
                    cell.font = total_font
                    cell.fill = total_fill
                    cell.alignment = Alignment(horizontal="right" if is_num else "center")
                else:
                    cell.alignment = Alignment(horizontal="right" if is_num else "left")
                if is_num:
                    cell.number_format = '"$"#,##0.00'

        # Auto-width columns
        for col in ws.columns:
            letter = col[0].column_letter
            width  = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[letter].width = max(width + 3, 14)

        ws.freeze_panes = "A2"

    wb = openpyxl.Workbook()

    # First sheet: All records
    ws_all = wb.active
    ws_all.title = "All"
    ws_all.sheet_properties.tabColor = "1D4ED8"
    write_sheet(ws_all, data_rows)

    # One sheet per recon status (only if it has data)
    # Exclude the TOTAL row from per-status sheets; add a fresh TOTAL row at the end
    non_total = [r for r in data_rows if str(r[0] or "").upper() != "TOTAL"]
    for status_val, sheet_name, tab_color in SHEETS:
        subset = [r for r in non_total if str(r[1] or "") == status_val]
        if not subset:
            continue
        # Build a totals row for this subset: sum all numeric columns
        totals_row = []
        for c_idx in range(len(header)):
            if c_idx == 0:
                totals_row.append("TOTAL")
            else:
                col_vals = [r[c_idx] for r in subset if c_idx < len(r)]
                nums = [v for v in col_vals if isinstance(v, (int, float))]
                totals_row.append(sum(nums) if nums else "")
        ws = wb.create_sheet(sheet_name)
        write_sheet(ws, subset + [totals_row], tab_color)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return {
        "data": base64.b64encode(buf.read()).decode("utf-8"),
        "filename": filename,
    }


@frappe.whitelist()
def setup_ar_custom_fields():
    """Create / update custom fields required by the AR dashboard.
    Call once after deploying: bench --site <site> call
    cannabis_management.cannabis_management.page.ar_dashboard.ar_dashboard.setup_ar_custom_fields
    """
    if not _can_edit_recon():
        frappe.throw("Administrator access required.", frappe.PermissionError)

    recon_options = "\n".join([""] + list(ALLOWED_RECON_STATUSES[1:]))

    # Update existing reconciliation status field options
    existing = frappe.db.get_value(
        "Custom Field",
        {"dt": "Customer", "fieldname": "custom_reconciliation_status"},
        "name",
    )
    if existing:
        frappe.db.set_value("Custom Field", existing, "options", recon_options)
        frappe.db.set_value("Custom Field", existing, "fieldtype", "Select")
    else:
        frappe.get_doc({
            "doctype": "Custom Field",
            "dt": "Customer",
            "fieldname": "custom_reconciliation_status",
            "label": "Reconciliation Status",
            "fieldtype": "Select",
            "options": recon_options,
            "insert_after": "customer_name",
        }).insert(ignore_permissions=True)

    # Onboarding / Company Made Contact — the two inline Yes/No columns.
    # These replaced the old POC column on the dashboard. custom_poc itself is
    # deliberately left in place: 108 customers still carry a Company/Nikki
    # value there and it means something different, so it is not reused.
    flag_options = "\n".join(ALLOWED_FLAG_VALUES)
    flag_labels = {
        "custom_onboarding": "Onboarding",
        "custom_company_made_contact": "Company Made Contact",
    }
    previous = "custom_poc"
    for fieldname, label in flag_labels.items():
        existing_flag = frappe.db.get_value(
            "Custom Field", {"dt": "Customer", "fieldname": fieldname}, "name"
        )
        if existing_flag:
            frappe.db.set_value("Custom Field", existing_flag, {
                "label": label,
                "fieldtype": "Select",
                "options": flag_options,
            })
        else:
            frappe.get_doc({
                "doctype": "Custom Field",
                "dt": "Customer",
                "fieldname": fieldname,
                "label": label,
                "fieldtype": "Select",
                "options": flag_options,
                "insert_after": previous,
            }).insert(ignore_permissions=True)
        previous = fieldname

    # New AR Available flag
    if not frappe.db.exists("Custom Field", {"dt": "Customer", "fieldname": "custom_new_ar_available"}):
        frappe.get_doc({
            "doctype": "Custom Field",
            "dt": "Customer",
            "fieldname": "custom_new_ar_available",
            "label": "New AR Available",
            "fieldtype": "Check",
            "default": "0",
            "insert_after": "custom_company_made_contact",
        }).insert(ignore_permissions=True)

    frappe.db.commit()
    return {"status": "ok", "message": "AR custom fields created/updated successfully."}
