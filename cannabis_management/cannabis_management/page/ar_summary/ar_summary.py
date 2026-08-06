"""Accounts Receivable — summary + per-customer ledgers.

Intercompany / internal-transfer invoices (is_internal_customer = 1, which Frappe
shows with the "Internal Transfer" status) are excluded from every view and export.

The ledger carries the Sales Order and Delivery Note the invoice was created against.
Outstanding balances are shown in red (page and export). The Excel export carries two
sheets: "Accounts Receivable" (summary) and "Customer Ledgers" (boxed blocks).
"""

import io

import frappe
from frappe.utils import flt, formatdate, get_url, nowdate

# ---- summary sheet colours -------------------------------------------------
YELLOW = "FFE699"   # Invoice Total
GREEN = "C6EFCE"    # Paid / settled outstanding
RED = "FFC7CE"      # outstanding still owed
HEADER = "DDEBF7"   # header row
RED_FONT = "C00000" # red text for outstanding balances

# ---- ledger sheet colours (match the Customer Ledgers screenshot) ----------
NAVY = "1F3864"     # per-customer title band
COLHDR = "D9D9D9"   # column-header row
AMT = "E2EFDA"      # amount cells in data rows (light green)
TOTAL_YELLOW = "FFE699"


def _conditions(company=None, from_date=None, to_date=None):
    """Shared WHERE clause. Always drops internal / intercompany transfers."""
    conditions = ["docstatus = 1", "IFNULL(is_internal_customer, 0) = 0"]
    values = {}
    if company:
        conditions.append("company = %(company)s")
        values["company"] = company
    if from_date:
        conditions.append("posting_date >= %(from_date)s")
        values["from_date"] = from_date
    if to_date:
        conditions.append("posting_date <= %(to_date)s")
        values["to_date"] = to_date
    return " AND ".join(conditions), values


def _summary_rows(company=None, from_date=None, to_date=None):
    where, values = _conditions(company, from_date, to_date)
    return frappe.db.sql(
        f"""
        SELECT customer,
               COUNT(*) AS invoices,
               SUM(grand_total) AS invoice_total,
               SUM(grand_total - outstanding_amount) AS paid,
               SUM(outstanding_amount) AS outstanding
        FROM `tabSales Invoice`
        WHERE {where}
        GROUP BY customer
        HAVING invoice_total > 0
        ORDER BY outstanding DESC, invoice_total DESC
        """,
        values,
        as_dict=True,
    )


def _attach_references(invoices):
    """Attach each invoice's Sales Order(s) — the SO it was created against, read
    from the invoice item rows — and the Delivery Note(s) raised against those
    Sales Orders (NOT the delivery_note stamped on the invoice item)."""
    names = [i["name"] for i in invoices]
    if not names:
        return

    # 1. Sales Order(s) each invoice was created against.
    rows = frappe.db.sql(
        """
        SELECT parent, sales_order
        FROM `tabSales Invoice Item`
        WHERE parent IN %(names)s AND IFNULL(sales_order, '') != ''
        """,
        {"names": tuple(names)},
        as_dict=True,
    )
    so_map, all_sos = {}, set()
    for r in rows:
        so_map.setdefault(r.parent, set()).add(r.sales_order)
        all_sos.add(r.sales_order)

    # 2. Delivery Note(s) raised against those Sales Orders.
    so_to_dn = {}
    if all_sos:
        dn_rows = frappe.db.sql(
            """
            SELECT against_sales_order AS so, parent AS dn
            FROM `tabDelivery Note Item`
            WHERE against_sales_order IN %(sos)s AND docstatus = 1
            """,
            {"sos": tuple(all_sos)},
            as_dict=True,
        )
        for r in dn_rows:
            so_to_dn.setdefault(r.so, set()).add(r.dn)

    # 3. Payment Entry(ies) made against each invoice, or against its Sales Order
    #    (advances). Read from Payment Entry Reference; submitted payments only.
    inv_pe, so_pe = {}, {}
    pe_rows = frappe.db.sql(
        """
        SELECT per.reference_doctype AS rdt, per.reference_name AS rname, per.parent AS pe
        FROM `tabPayment Entry Reference` per
        INNER JOIN `tabPayment Entry` pe ON pe.name = per.parent AND pe.docstatus = 1
        WHERE (per.reference_doctype = 'Sales Invoice' AND per.reference_name IN %(inv)s)
           OR (per.reference_doctype = 'Sales Order' AND per.reference_name IN %(sos)s)
        """,
        {"inv": tuple(names), "sos": tuple(all_sos) if all_sos else ("",)},
        as_dict=True,
    )
    for r in pe_rows:
        (inv_pe if r.rdt == "Sales Invoice" else so_pe).setdefault(r.rname, set()).add(r.pe)

    for inv in invoices:
        sos = so_map.get(inv["name"], set())
        inv["sales_order"] = sorted(sos)
        dns = set()
        for so in sos:
            dns |= so_to_dn.get(so, set())
        pes = set(inv_pe.get(inv["name"], set()))
        for so in sos:
            pes |= so_pe.get(so, set())
        inv["payment_entries"] = sorted(pes)
        inv["delivery_note"] = sorted(dns)


def _ledger_groups(company=None, from_date=None, to_date=None):
    """Per-invoice detail grouped by customer, groups sorted by outstanding
    balance (highest first), matching the reference sheet."""
    where, values = _conditions(company, from_date, to_date)
    invoices = frappe.db.sql(
        f"""
        SELECT name, customer, company, posting_date, due_date,
               grand_total AS invoice_total,
               (grand_total - outstanding_amount) AS paid,
               outstanding_amount AS outstanding,
               status
        FROM `tabSales Invoice`
        WHERE {where}
        ORDER BY customer, posting_date, name
        """,
        values,
        as_dict=True,
    )
    _attach_references(invoices)

    groups = {}
    for inv in invoices:
        groups.setdefault(inv.customer, []).append(inv)

    result = []
    for customer, invs in groups.items():
        totals = {
            "invoice_total": sum(flt(i.invoice_total) for i in invs),
            "paid": sum(flt(i.paid) for i in invs),
            "outstanding": sum(flt(i.outstanding) for i in invs),
        }
        if totals["invoice_total"] <= 0:  # keep in step with the summary
            continue
        result.append({
            "customer": customer,
            "invoices": invs,
            "count": len(invs),
            "totals": totals,
        })

    result.sort(key=lambda g: g["totals"]["outstanding"], reverse=True)
    return result


@frappe.whitelist()
def get_ar_summary(company=None, from_date=None, to_date=None):
    rows = _summary_rows(company, from_date, to_date)
    totals = {
        "invoices": sum(r.invoices for r in rows),
        "invoice_total": sum(flt(r.invoice_total) for r in rows),
        "paid": sum(flt(r.paid) for r in rows),
        "outstanding": sum(flt(r.outstanding) for r in rows),
    }
    return {
        "rows": rows,
        "totals": totals,
        "prepared": formatdate(nowdate(), "MMMM dd, yyyy"),
    }


@frappe.whitelist()
def get_ledgers(company=None, from_date=None, to_date=None):
    groups = _ledger_groups(company, from_date, to_date)
    for g in groups:
        for inv in g["invoices"]:
            inv["posting_date"] = formatdate(inv.posting_date, "yyyy-MM-dd") if inv.posting_date else ""
            inv["due_date"] = formatdate(inv.due_date, "yyyy-MM-dd") if inv.due_date else ""
    return {
        "groups": groups,
        "prepared": formatdate(nowdate(), "MMMM dd, yyyy"),
    }


# ===========================================================================
# Excel export — two sheets
# ===========================================================================

def _build_summary_sheet(ws, rows, prepared):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    money = "$#,##0.00"
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")

    def fill(hexcode):
        return PatternFill("solid", fgColor=hexcode)

    ws.title = "Accounts Receivable"
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Prepared {prepared} — Accounts Receivable"
    ws["A1"].font = Font(italic=True, size=10, color="595959")

    headers = ["Customer", "# Invoices", "Invoice Total", "Paid Amount", "Outstanding"]
    hr = 3
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = fill(HEADER)
        cell.border = border
        cell.alignment = center if c > 1 else Alignment(horizontal="left")

    r = hr + 1
    for row in rows:
        outstanding = flt(row.outstanding)
        paid = flt(row.paid)
        cust = ws.cell(row=r, column=1, value=row.customer)
        inv = ws.cell(row=r, column=2, value=row.invoices)
        it = ws.cell(row=r, column=3, value=flt(row.invoice_total))
        pd = ws.cell(row=r, column=4, value=paid)
        out = ws.cell(row=r, column=5, value=outstanding)

        for cell in (cust, inv, it, pd, out):
            cell.border = border
        inv.alignment = center
        for cell in (it, pd, out):
            cell.number_format = money
            cell.alignment = right

        it.fill = fill(YELLOW)
        if paid > 0:
            pd.fill = fill(GREEN)
        # outstanding owed -> red; settled -> green
        if outstanding > 0:
            out.fill = fill(RED)
            out.font = Font(color=RED_FONT, bold=True)
        else:
            out.fill = fill(GREEN)
        r += 1

    tr = r
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=tr, column=2, value=sum(x.invoices for x in rows)).alignment = center
    for col, key in ((3, "invoice_total"), (4, "paid"), (5, "outstanding")):
        cell = ws.cell(row=tr, column=col, value=sum(flt(x[key]) for x in rows))
        cell.number_format = money
        cell.font = Font(bold=True, color=RED_FONT) if col == 5 else Font(bold=True)
        cell.alignment = right
    for c in range(1, 6):
        ws.cell(row=tr, column=c).border = border

    for col, w in {"A": 34, "B": 11, "C": 16, "D": 16, "E": 16}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"


def _ref_cell(ws, row, col, names, doctype, border, fill_amt, left):
    """Write a Sales Order / Delivery Note reference cell. Hyperlinks a single
    reference; comma-joins multiples as plain text."""
    from openpyxl.styles import Font

    names = names or []
    cell = ws.cell(row=row, column=col, value=", ".join(names))
    cell.border = border
    cell.alignment = left
    cell.fill = fill_amt
    if len(names) == 1:
        cell.hyperlink = f"{get_url()}/app/{doctype}/{names[0]}"
        cell.font = Font(color="0563C1", underline="single")
    return cell


def _build_ledger_sheet(ws, groups, prepared):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    money = "$#,##0.00"
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")
    left = Alignment(horizontal="left")

    def fill(hexcode):
        return PatternFill("solid", fgColor=hexcode)

    fill_amt = fill(AMT)
    fill_ref = fill(AMT)  # light tint behind reference cells too (kept subtle/white-ish)

    ws.title = "Customer Ledgers"
    LAST = 10  # A:J
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=LAST)
    ws["A1"] = (f"Prepared {prepared} — each customer is its own boxed block, "
                "sorted by outstanding balance (highest first)")
    ws["A1"].font = Font(italic=True, size=10, color="595959")

    headers = ["Invoice #", "Sales Order", "Payment",
               "Company", "Posting Date", "Due Date", "Invoice Total",
               "Paid Amount", "Outstanding", "Status"]

    r = 3
    for g in groups:
        n = g["count"]
        label = f"{g['customer']}   ({n} invoice{'s' if n != 1 else ''})"

        # customer title band (navy, white, merged A:J)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=LAST)
        band = ws.cell(row=r, column=1, value=label)
        band.font = Font(bold=True, color="FFFFFF", size=11)
        band.fill = fill(NAVY)
        band.alignment = left
        for c in range(1, LAST + 1):
            ws.cell(row=r, column=c).border = border
        r += 1

        # column headers
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = Font(bold=True)
            cell.fill = fill(COLHDR)
            cell.border = border
            cell.alignment = center if c >= 5 else left
        r += 1

        # invoice rows
        for inv in g["invoices"]:
            name_cell = ws.cell(row=r, column=1, value=inv["name"])
            name_cell.hyperlink = f"{get_url()}/app/sales-invoice/{inv['name']}"
            name_cell.font = Font(color="0563C1", underline="single")
            name_cell.alignment = left
            name_cell.border = border

            _ref_cell(ws, r, 2, inv.get("sales_order"), "sales-order", border, PatternFill(), left)
            _ref_cell(ws, r, 3, inv.get("payment_entries"), "payment-entry", border, PatternFill(), left)

            ws.cell(row=r, column=4, value=inv["company"]).alignment = left
            ws.cell(row=r, column=5,
                    value=formatdate(inv["posting_date"], "yyyy-MM-dd") if inv["posting_date"] else "").alignment = center
            ws.cell(row=r, column=6,
                    value=formatdate(inv["due_date"], "yyyy-MM-dd") if inv["due_date"] else "").alignment = center

            it = ws.cell(row=r, column=7, value=flt(inv["invoice_total"]))
            pd = ws.cell(row=r, column=8, value=flt(inv["paid"]))
            out = ws.cell(row=r, column=9, value=flt(inv["outstanding"]))
            for cell in (it, pd):
                cell.number_format = money
                cell.alignment = right
                cell.fill = fill_amt
            out.number_format = money
            out.alignment = right
            # outstanding owed -> red; settled -> green
            if flt(inv["outstanding"]) > 0:
                out.fill = fill(RED)
                out.font = Font(color=RED_FONT, bold=True)
            else:
                out.fill = fill(GREEN)

            ws.cell(row=r, column=10, value=inv["status"]).alignment = center

            for c in range(1, LAST + 1):
                ws.cell(row=r, column=c).border = border
            r += 1

        # per-customer TOTAL row (yellow); outstanding total in red
        t = g["totals"]
        ws.cell(row=r, column=1, value="TOTAL")
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = fill(TOTAL_YELLOW)
            ws.cell(row=r, column=c).font = Font(bold=True)
        for col, key in ((7, "invoice_total"), (8, "paid"), (9, "outstanding")):
            cell = ws.cell(row=r, column=col, value=flt(t[key]))
            cell.number_format = money
            cell.alignment = right
            cell.fill = fill(TOTAL_YELLOW)
            cell.font = Font(bold=True, color=RED_FONT) if col == 9 else Font(bold=True)
        ws.cell(row=r, column=10).fill = fill(TOTAL_YELLOW)
        for c in range(1, LAST + 1):
            ws.cell(row=r, column=c).border = border
        r += 2  # blank spacer between customers

    widths = {"A": 22, "B": 20, "C": 20, "D": 24, "E": 13,
              "F": 13, "G": 14, "H": 14, "I": 14, "J": 13}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"


@frappe.whitelist()
def export_ar_summary(company=None, from_date=None, to_date=None):
    from openpyxl import Workbook

    rows = _summary_rows(company, from_date, to_date)
    groups = _ledger_groups(company, from_date, to_date)
    prepared = formatdate(nowdate(), "MMMM dd, yyyy")

    wb = Workbook()
    _build_summary_sheet(wb.active, rows, prepared)
    _build_ledger_sheet(wb.create_sheet(), groups, prepared)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    frappe.response["filename"] = f"Accounts Receivable {prepared}.xlsx"
    frappe.response["filecontent"] = bio.read()
    frappe.response["type"] = "binary"